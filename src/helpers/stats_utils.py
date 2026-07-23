import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.stats import norm, gaussian_kde
from openpyxl import Workbook
from src.helpers import plot_utils
# remove later
from src.helpers import line_utils, data_utils, matrix_utils, plot_utils
from src.algorithms import classifier, estimators, reconstructor

def bootstrap_by_gaps(
    residuals,
    uncertainties,
    column_indices,
    interp_mask=None,
    n_boot=500,
    ci=95,
    min_block_width=2,
    fraction=0.1
):
    abs_res = np.abs(residuals)
    unc = uncertainties
    valid = np.isfinite(abs_res) & np.isfinite(unc)

    if interp_mask is not None:
        valid = valid & (~interp_mask)

    column_indices = np.asarray(column_indices)

    # --- Build candidate blocks (consecutive gaps) ---
    candidate_blocks = []
    for i in range(len(column_indices) - 1):
        start = column_indices[i]
        end = column_indices[i + 1]
        if end - start >= min_block_width:
            candidate_blocks.append((start, end))

    if len(candidate_blocks) == 0:
        raise ValueError("No valid blocks found.")

    # --- Compute max_gaps_per_block from physical size ---
    if fraction is not None:
        n_cols = residuals.shape[1]
        max_gaps_per_block = max(1, int(n_cols * fraction))
    else:
        max_gaps_per_block = 20  # default fallback

    p_boot = []

    for _ in range(n_boot):
        start_idx = np.random.randint(0, len(candidate_blocks))
        n_gaps = np.random.randint(1, max_gaps_per_block + 1)
        end_idx = min(start_idx + n_gaps, len(candidate_blocks))

        c0 = candidate_blocks[start_idx][0]
        c1 = candidate_blocks[end_idx - 1][1]

        r_block = abs_res[:, c0:c1]
        u_block = unc[:, c0:c1]
        m_block = valid[:, c0:c1]

        if np.any(m_block):
            indicators = u_block[m_block] >= r_block[m_block]
            p_boot.append(np.mean(indicators))

    p_boot = np.array(p_boot) * 100
    alpha = 100 - ci

    ci_stats = {
        "mean": np.mean(p_boot),
        "ci_lower": np.percentile(p_boot, alpha/2),
        "ci_upper": np.percentile(p_boot, 100-alpha/2),
        "std": np.std(p_boot),
        "boot": p_boot
    }

    return ci_stats

def uncertainty_comparison_for_multi(residuals, uncertainties, interp_mask, thresholds, eps=1e-12):

    if residuals.ndim < 2:
        residuals = residuals.reshape(1,-1)
    if uncertainties.ndim < 2:
        uncertainties = uncertainties.reshape(1, -1)

    if interp_mask is None:
        interp_mask = np.zeros_like(residuals, dtype=bool)

    valid = (~interp_mask) & np.isfinite(residuals) & np.isfinite(uncertainties)

    abs_res = np.abs(residuals[valid])
    unc = uncertainties[valid]

    # -------------------------------------------------
    # UNCERTAINTY RATIO (coverage metric)
    # -------------------------------------------------
    ratio = unc / np.maximum(abs_res, eps)

    total_count = ratio.size
    fail_count = np.sum(ratio < 1)

    pass_percentage = 100 * (1 - fail_count / total_count)

    # -------------------------------------------------
    # ACCURACY METRICS
    # -------------------------------------------------
    diff = unc - abs_res

    rmse = np.sqrt(np.mean(diff ** 2))
    mae = np.mean(np.abs(diff))
    mean_error = np.mean(diff)
    std_dev = np.std(diff)

    # -------------------------------------------------
    # SHARPNESS (smaller uncertainty preferred)
    # -------------------------------------------------
    sharp = np.mean(unc)

    # -------------------------------------------------
    # CORRELATION (uncertainty vs error magnitude)
    # -------------------------------------------------
    if total_count > 1:
        corr = np.corrcoef(unc, abs_res)[0, 1]
    else:
        corr = np.nan

    # Tail distribution
    # --- Normalize ---
    z = (residuals / uncertainties)
    z = z[valid]
    z = z[np.isfinite(z)]
    z_flat = z.flatten()

    # --- Stats ---
    mean_z = np.mean(z_flat)
    var_z = np.var(z_flat)

    # --- Tails ---
    # --- Total count ---
    N = len(z_flat)

    # --- Gaussian references ---
    ref = {"Mean": 0.0, "Variance": 1.0, "Coverage_1sigma": 0.6827, "Coverage_2sigma": 0.9545, "Coverage_3sigma": 0.9973,}

    # Tail references
    ref_tails = {3: 0.0027, 5: 5.733e-7, 10: 7.62e-24}

    tails = {t: np.mean(np.abs(z_flat) > t) for t in thresholds}

    # --- Counts ---
    tail_counts = {t: int(np.sum(np.abs(z_flat) > t)) for t in thresholds}

    # --- Ratios vs Gaussian ---
    tail_ratios = {t: tails[t] / ref_tails[t] if ref_tails[t] > 0 else np.nan for t in thresholds}

    tail_stats = {
        "tails": tails,
        "ref_tails": ref_tails,
        "tail_ratios": tail_ratios,
        "tail_counts": tail_counts
    }

    # -------------------------------------------------
    # SEVERITY OF UNDER-ESTIMATION
    # -------------------------------------------------
    under_mask = unc < abs_res

    if np.any(under_mask):
        severity_abs = abs_res[under_mask] - unc[under_mask]
        severity_rel = severity_abs / np.maximum(abs_res[under_mask], eps)

        severity_mean = np.mean(severity_rel)
        severity_tail = np.percentile(severity_rel, 95)
    else:
        severity_mean = 0.0
        severity_tail = 0.0

    alpha = 0.8
    severity_score = (alpha * severity_tail) + ((1 - alpha) * severity_mean)

    kappa = 1.5  # allowable safety margin

    over_mask = unc > (kappa * abs_res)

    if np.any(over_mask):
        over_abs = unc[over_mask] - (kappa * abs_res[over_mask])
        over_rel = over_abs / np.maximum(unc[over_mask], eps)

        over_mean = np.mean(over_rel)
        over_tail = np.percentile(over_rel, 95)
    else:
        over_mean = 0.0
        over_tail = 0.0

    over_penalty = (alpha * over_tail) + ((1 - alpha) * over_mean)

    # -------------------------------------------------
    # ADDED: reliability score (calibration honesty)
    # -------------------------------------------------
    confidence_levels = [0.5, 0.68, 0.9, 0.95, 0.99]

    reliability_errors = []

    r = abs_res
    u = unc

    for p in confidence_levels:
        u_p = np.percentile(u, p * 100)
        empirical_cov = np.mean(r <= u_p)
        reliability_errors.append(abs(empirical_cov - p))

    reliability_error = np.mean(reliability_errors)

    return pass_percentage, rmse, mae, mean_error, std_dev, sharp, corr, severity_mean, severity_tail, severity_score, over_mean, over_tail, over_penalty, reliability_error, tail_stats


def multi_uncertainty_comparison(
    residuals,
    uncertainties_dict,
    resolution,
    desired_linespacing_meters=None,
    column_indices = None,
    seabed=None,
    plot_grid=(3, 3),
    path=None,
    plot_boxplots=True,
    plot_reliability=True,
    thresholds = {3,5,10},
    confidence_levels = [0.5, 0.68, 0.9, 0.95, 0.99],
    alpha = 0.8, # for computing severity score
    normalize_residual = True,
    interp_mask=None,
    bootstrap = True,
):
    """
    Compare multiple uncertainty surfaces against residuals in one figure.

    Parameters
    ----------
    residuals : np.ndarray
        2D array of residual surface.
    uncertainties_dict : dict
        Dictionary of uncertainty name -> uncertainty array.
    resolution : float
        Grid resolution in meters.
    desired_linespacing_meters : float, optional
        Used for labeling titles.
    fn : str, optional
        Surface name for the first title.
    plot_grid : tuple
        (nrows, ncols) for subplot grid.
    """

    # ---- Create figure ----
    nrows, ncols = plot_grid
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(12, 8), layout="constrained")

    axes = axes.flatten()
    names = list(uncertainties_dict.keys())
    results = []

    for i, (name, uncertainty) in enumerate(uncertainties_dict.items()):
        ax = axes[i]

        # Compute stats
        (pass_percentage, rmse, mae, mean_error, std_dev, sharp, corr, severity_mean,
         severity_tail, severity_score, over_mean, over_tail, over_penalty,
         reliability_error, tail_stats) = uncertainty_comparison_for_multi(residuals, uncertainty, interp_mask=interp_mask, thresholds=thresholds)

        # Append stats for CSV
        results.append({
            "Seabed": seabed,
            "Uncertainty Method": name,
            "Line spacing ": desired_linespacing_meters,
            "Pass %": pass_percentage,
            "RMSE": rmse,
            "MAE": mae,
            "Bias (Mean Error)": mean_error,
            "Std Dev": std_dev,
            "Sharpness": sharp,
            "Correlation": corr,
            "Severity_mean": severity_mean,
            "Severity_tail": severity_tail,
            "Severity score": severity_score,
            "Over_mean": over_mean,
            "Over_tail": over_tail,
            "Over_penalty": over_penalty,
            "Reliability score": reliability_error
        })

        if bootstrap:
            ci_stats = bootstrap_by_gaps(residuals, uncertainty, column_indices, n_boot=500, ci=95,
                                         interp_mask=interp_mask, min_block_width=2, fraction=0.6)  # 500
            # results["ci_stats"] = ci_stats
            results[-1].update({"CI mean": ci_stats["mean"], "CI lower": ci_stats["ci_lower"], "CI upper": ci_stats["ci_upper"], "CI std": ci_stats["std"]})

        for t in thresholds:
            results[-1].update({
            f"P_gt_{t}": tail_stats['tails'][t],
            f"Gaussian_{t}": tail_stats['ref_tails'][t],
            f"Ratio_{t}": tail_stats['tail_ratios'][t],
            f"Count_{t}": tail_stats['tail_counts'][t],
        })

        # Scatter comparison plot
        nonzero_idx = np.nonzero(
            (residuals != 0) & (~np.isnan(residuals)) & (uncertainty != 0)
        )
        max_unc = np.max(uncertainty[nonzero_idx])
        ax.plot(np.abs(residuals[nonzero_idx]), uncertainty[nonzero_idx], ".", alpha=0.3)
        # ax.scatter(np.abs(residuals[nonzero_idx]),uncertainty[nonzero_idx],s=1,alpha=0.15,rasterized=True, linewidths=0)
        # ax.hexbin(np.abs(residuals[nonzero_idx]), uncertainty[nonzero_idx], gridsize=60, bins='log', mincnt=5, cmap='magma')
        ax.plot([0, max_unc], [0, max_unc], "r", lw=1)

        ax.set_xlabel("Abs. Residual (% of depth)") if normalize_residual == True else ax.set_xlabel("Abs. Residual (m)")
        ax.set_ylabel("Uncertainty (% of depth)") if normalize_residual == True else ax.set_ylabel("Uncertainty (m)")
        ax.set_xlim(0, max_unc)
        ax.set_ylim(0, max_unc)
        ax.grid(True, alpha=0.3)

        # Title with stats
        ax.set_title(
            f"{name}\nPass: {ci_stats['mean'] if bootstrap else pass_percentage:.1f}%  RMSE: {rmse:.2f}  Bias: {mean_error:.2f}\n  "
            f"Severity mean: {severity_mean:.2f} Severity: {severity_score:.2f}\n"
            f"Over_penalty: {over_penalty:.2f} MAE: {mae:.2f} Corr: {corr:.2f}", fontsize=14)

    # Remove unused axes
    for j in range(i + 1, len(axes)):
        fig.delaxes(axes[j])

    if seabed and desired_linespacing_meters:
        fig.suptitle(
            f"Uncertainty Comparisons for {seabed} ({resolution}m grid, {desired_linespacing_meters}m spacing)",
            fontsize=14,
        )
    else:
        fig.suptitle("Uncertainty Comparisons", fontsize=14)
    outpath = f'{path}_uncertainty_comparisons.png'
    plt.savefig(outpath, bbox_inches='tight')
    plt.show()


    # ---- Export to CSV ----
    if path:
        df = pd.DataFrame(results)
        outpath = f'{path}_stats.csv'
        df.to_csv(f'{outpath}', index=False)
        print(f"Statistics exported to {outpath}")

    # ---- Optional: Combined Boxplot of Residuals vs Uncertainties ----

    if plot_boxplots:
        # Collect data
        data = []
        labels = []

        # Residuals (absolute values)
        res_vals = np.abs(residuals[(residuals != 0) & (~np.isnan(residuals))]).flatten()
        data.append(res_vals)
        labels.append("N. Abs. Residuals") if normalize_residual == True else labels.append("Abs. Residuals")

        # Each uncertainty
        for name, uncertainty in uncertainties_dict.items():
            unc_vals = uncertainty[(uncertainty != 0) & (~np.isnan(uncertainty))].flatten()
            data.append(unc_vals)
            labels.append(name)

        # Combined boxplot
        plt.figure(figsize=(10, 8))
        plt.boxplot(data, patch_artist=True, labels=labels,
                    boxprops=dict(facecolor='lightgray', alpha=0.7),
                    medianprops=dict(color='red', linewidth=1.5))
        # plt.title(f"Uncertainty Boxplots for {fn} ({resolution}m grid, {desired_linespacing_meters}m spacing)")
        plt.ylabel("Uncertainty (% of depth)") if normalize_residual == True else plt.ylabel("Uncertainty (m)", fontsize=14)
        plt.grid(alpha=0.3)
        plt.xticks(rotation=30, fontsize=14)
        plt.yticks(fontsize=14)

        outpath = f'{path}_uncertainty_boxplots.png'
        plt.savefig(outpath, bbox_inches='tight')
        plt.show()

    if plot_reliability:

        residuals = np.abs(residuals)
        mask = ~np.isnan(residuals)
        reliability_results = {}

        plt.figure(figsize=(10, 8))

        for name, unc in uncertainties_dict.items():
            unc = np.abs(unc)
            m = mask & ~np.isnan(unc)
            r = residuals[m]
            u = unc[m]

            empirical_cov = []
            for p in confidence_levels:
                # Determine the uncertainty threshold corresponding to the p-th percentile
                unc_threshold = np.percentile(u, p * 100)
                # Fraction of residuals covered by that threshold
                coverage = np.mean(r <= unc_threshold)
                empirical_cov.append(coverage)

            reliability_results[name] = np.array(empirical_cov)
            plt.plot(confidence_levels, empirical_cov, marker='o', label=name)

        # Ideal calibration line
        plt.plot([0.5, 1], [0.5, 1], 'k--', lw=1.2, label='Perfect calibration')

        plt.xlabel('Nominal Confidence Level')
        plt.ylabel('Empirical Coverage')
        plt.title(f'Nonparametric Reliability Curves for {seabed} ({resolution}m grid, {desired_linespacing_meters}m spacing)')
        plt.grid(alpha=0.3)
        plt.legend()
        plt.tight_layout()
        outpath = f'{path}_uncertainty_reliability_diagram.png'
        plt.savefig(outpath, bbox_inches='tight')
        plt.show()
    # return results


def uncertainty_comparison(residuals, uncertainties, interp_mask=None, eps=1e-12):

    # -------------------------------------------------
    # VALID MASK
    # -------------------------------------------------
    if interp_mask is None:
        interp_mask = np.zeros_like(residuals, dtype=bool)

    # valid = (~interp_mask) & np.isfinite(residuals) & np.isfinite(uncertainties) & (residuals != 0)
    # valid = (residuals != 0) & np.isfinite(residuals) & np.isfinite(uncertainties)
    # valid = (residuals != 0) & (uncertainties != 0) & np.isfinite(residuals) & np.isfinite(uncertainties)
    valid = (~interp_mask) & np.isfinite(residuals) & np.isfinite(uncertainties)

    abs_res = np.abs(residuals[valid])
    unc = uncertainties[valid]


    # Tail distribution
    # --- Normalize ---
    z = (residuals / uncertainties)
    z = z[valid]
    z = z[np.isfinite(z)]
    z_flat = z.flatten()

    # --- Stats ---
    mean_z = np.mean(z_flat)
    var_z = np.var(z_flat)
    # alpha = np.sqrt(np.mean(z_flat ** 2))
    # cov1 = np.mean(np.abs(z_flat) <= 1)
    # cov2 = np.mean(np.abs(z_flat) <= 2)
    # cov3 = np.mean(np.abs(z_flat) <= 3)

    # --- Tails ---
    # --- Total count ---
    N = len(z_flat)

    # --- Gaussian references ---
    ref = {
        "Mean": 0.0,
        "Variance": 1.0,
        "Coverage_1sigma": 0.6827,
        "Coverage_2sigma": 0.9545,
        "Coverage_3sigma": 0.9973,
    }

    # threshold
    thresholds = (3, 5, 10)

    # Tail references
    ref_tails = {3: 0.0027, 5: 5.733e-7, 10: 7.62e-24  # effectively zero
                 }

    tails = {t: np.mean(np.abs(z_flat) > t) for t in thresholds}

    # --- Counts ---
    tail_counts = {t: int(np.sum(np.abs(z_flat) > t)) for t in thresholds}

    # --- Ratios vs Gaussian ---
    tail_ratios = {
        t: tails[t] / ref_tails[t] if ref_tails[t] > 0 else np.nan
        for t in thresholds
    }

    # -------------------------------------------------
    # UNCERTAINTY RATIO (coverage metric)
    # -------------------------------------------------
    ratio = unc / np.maximum(abs_res, eps)

    total_count = ratio.size
    fail_count = np.sum(ratio < 1)

    pass_percentage = 100 * (1 - fail_count / total_count)

    # -------------------------------------------------
    # ACCURACY METRICS
    # -------------------------------------------------
    diff = unc - abs_res

    rmse = np.sqrt(np.mean(diff**2))
    mae = np.mean(np.abs(diff))
    mean_error = np.mean(diff)
    std_dev = np.std(diff)

    # -------------------------------------------------
    # SHARPNESS (smaller uncertainty preferred)
    # -------------------------------------------------
    sharp = np.mean(unc)

    # -------------------------------------------------
    # CORRELATION (uncertainty vs error magnitude)
    # -------------------------------------------------
    if total_count > 1:
        corr = np.corrcoef(unc, abs_res)[0, 1]
    else:
        corr = np.nan

    # -------------------------------------------------
    # SEVERITY OF UNDER-ESTIMATION
    # -------------------------------------------------
    under_mask = unc < abs_res

    if np.any(under_mask):
        severity_abs = abs_res[under_mask] - unc[under_mask]
        severity_rel = severity_abs / np.maximum(abs_res[under_mask], eps)

        severity_mean = np.mean(severity_rel)
        severity_tail = np.percentile(severity_rel, 95)
    else:
        severity_mean = 0.0
        severity_tail = 0.0

    alpha = 0.8
    severity_score = (alpha * severity_tail) + ((1 - alpha) * severity_mean)

    kappa = 1.5  # allowable safety margin

    over_mask = unc > (kappa * abs_res)

    if np.any(over_mask):
        over_abs = unc[over_mask] - (kappa * abs_res[over_mask])
        over_rel = over_abs / np.maximum(unc[over_mask], eps)

        over_mean = np.mean(over_rel)
        over_tail = np.percentile(over_rel, 95)
    else:
        over_mean = 0.0
        over_tail = 0.0

    over_penalty = (alpha * over_tail) + ((1 - alpha) * over_mean)

    # -------------------------------------------------
    # ADDED: reliability score (calibration honesty)
    # -------------------------------------------------
    confidence_levels = [0.5, 0.68, 0.9, 0.95, 0.99]

    reliability_errors = []

    r = abs_res
    u = unc

    for p in confidence_levels:
        u_p = np.percentile(u, p * 100)
        empirical_cov = np.mean(r <= u_p)
        reliability_errors.append(abs(empirical_cov - p))

    reliability_error = np.mean(reliability_errors)

    # -------------------------------------------------
    # RETURN RESULTS
    # -------------------------------------------------
    results = {
        "total_cts": total_count,
        "fail_cts": fail_count,
        "percentage": pass_percentage,
        "rmse": rmse,
        "mae": mae,
        "mean": mean_error,
        "std_dev": std_dev,
        "sharp": sharp,
        "corr": corr,
        "severity_mean": severity_mean,
        "severity_tail": severity_tail,
        "severity_score": severity_score,
        "over_mean": over_mean,
        "over_tail": over_tail,
        "over_penalty": over_penalty,
        "reliability_error": reliability_error,
        "mean_norm_res": mean_z,
        "var_norm_res": var_z,
        "Total_N": N}

    for t in thresholds:
        results.update({
            f"P_gt_{t}": tails[t],
            f"Gaussian_{t}": ref_tails[t],
            f"Ratio_{t}": tail_ratios[t],
            f"Count_{t}": tail_counts[t],
        })

    return results, unc, abs_res


def evaluate_uncertainty_model(residuals, uncertainty, label, outdir, extent, ndv, xsize, ysize, res, seabed_name, spacing, interp_mask = None, thresholds=(3, 5), clip_ranges=None, plot=True,
):


    """
    Uncertainty evaluation with:
    - Calibration (mean, var, alpha)
    - Reliability (coverage)
    - Tail behavior
    - Asymmetry
    - PNG plots
    - XLSX stats export
    """

    if clip_ranges is None:
        clip_ranges = [(-3, 3), (-5, 5)]

    os.makedirs(outdir, exist_ok=True)

    if interp_mask is not None:
        valid = (~interp_mask) & np.isfinite(residuals) & np.isfinite(uncertainty)
    else:
        valid = np.isfinite(residuals) & np.isfinite(uncertainty)

    z = residuals / uncertainty

    # preserve shape
    z = np.where(valid & np.isfinite(z), z, np.nan)
    z_flat = z[~np.isnan(z)]

    # --- Stats ---
    mean_z = np.mean(z_flat)
    var_z = np.var(z_flat)
    alpha = np.sqrt(np.mean(z_flat**2))

    cov1 = np.mean(np.abs(z_flat) <= 1)
    cov2 = np.mean(np.abs(z_flat) <= 2)
    cov3 = np.mean(np.abs(z_flat) <= 3)

    # --- Tails ---
    # --- Total count ---
    N = len(z_flat)

    # --- Gaussian references ---
    ref = {"Mean": 0.0, "Variance": 1.0, "Coverage_1sigma": 0.6827, "Coverage_2sigma": 0.9545, "Coverage_3sigma": 0.9973}

    # Tail references
    ref_tails = {3: 0.0027,5: 5.733e-7,10: 7.62e-24}

    tails = {t: np.mean(np.abs(z_flat) > t) for t in thresholds}

    # --- Counts ---
    tail_counts = {t: int(np.sum(np.abs(z_flat) > t)) for t in thresholds}

    # --- Ratios vs Gaussian ---
    tail_ratios = {t: tails[t] / ref_tails[t] if ref_tails[t] > 0 else np.nan for t in thresholds}

    # --- Asymmetry ---
    pos = z_flat[z_flat > 0]
    neg = z_flat[z_flat < 0]

    pos_mean = np.mean(np.abs(pos)) if len(pos) else np.nan
    neg_mean = np.mean(np.abs(neg)) if len(neg) else np.nan

    print(f"\n===== {label} ({seabed_name}) =====")
    print(f"Mean: {mean_z:.4f}, Var: {var_z:.4f}, Alpha: {alpha:.4f}")
    print(f"Coverage: {cov1:.3f}, {cov2:.3f}, {cov3:.3f}")
    print(f"Tails: {tails}")
    print(f"Asymmetry: pos={pos_mean:.3f}, neg={neg_mean:.3f}")

    # =====================
    # Plot: PDF
    # =====================
    parameters = {'axes.labelsize': 14,
                'axes.titlesize': 14,
                'xtick.labelsize': 12,
                'ytick.labelsize': 12}
    plt.rcParams.update(parameters)
    if plot:
        for clip_range in clip_ranges:
            # plt.figure(figsize=(8, 6))
            plt.figure(figsize=(4.5,3), dpi=600, tight_layout=True)
            kde = gaussian_kde(z_flat)
            if clip_range is None:
                clip_range = (np.min(z), np.max(z))
            x = np.linspace(clip_range[0], clip_range[1], 500)

            plt.plot(x, kde(x), label=f"KDE")
            plt.plot(x, norm.pdf(x, 0, 1), '--', color='black', label="N(0,1)")

            plt.xlim(clip_range)
            plt.xlabel("Normalized Residual (z)")
            plt.ylabel("Probability Density")
            # plt.title(f"{label} ({seabed_name} {spacing}m)")
            plt.legend(fontsize=10)
            plt.grid(False)

            fname = os.path.join(outdir, f"{seabed_name}_{spacing}m_{label}_{clip_range[1]}_pdf.png")
            plt.savefig(fname, dpi=300, bbox_inches="tight")
            # plt.close()

    # =====================
    # Spatial maps
    # =====================
    # if plot and spatial_shape is not None:
    if plot:

        for t in thresholds:
            mask = (np.abs(z) > t).astype(float)

            plot_utils.plot_figure_mask(array=mask, extent=extent, ndv=ndv, outdir=outdir, xsize=xsize, ysize=ysize, res=res, seabed=seabed_name, spacing=spacing, title=f'{label}_z_gt_{t}', cb_label=t, scale_colour='black', utm_zone = None)

    # =====================
    # Save stats to Excel
    # =====================
    wb = Workbook()
    # =====================
    # Sheet 1: Summary
    # =====================
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Observed", "Reference"])

    ws.append(["Mean", mean_z, ref["Mean"]])
    ws.append(["Variance", var_z, ref["Variance"]])
    ws.append(["Alpha", alpha, "—"])

    ws.append(["Coverage_1sigma", cov1, ref["Coverage_1sigma"]])
    ws.append(["Coverage_2sigma", cov2, ref["Coverage_2sigma"]])
    ws.append(["Coverage_3sigma", cov3, ref["Coverage_3sigma"]])

    # =====================
    # Sheet 2: Tails
    # =====================
    ws2 = wb.create_sheet(title="Tails")
    ws2.append(["Threshold", "P(|r / σ|>t)", "Gaussian", "Ratio", "Count", "Total_N"])

    for t in thresholds:
        ws2.append([
            t,
            tails[t],
            ref_tails[t],
            tail_ratios[t],
            tail_counts[t],
            N
        ])

    # =====================
    # Sheet 3: Asymmetry
    # =====================
    ws3 = wb.create_sheet(title="Asymmetry")
    ws3.append(["Metric", "Value"])

    ws3.append(["Pos_Mean_Abs", pos_mean])
    ws3.append(["Neg_Mean_Abs", neg_mean])

    # Save
    xlsx_path = os.path.join(outdir, f"{seabed_name}_{spacing}m_{label}_stats.xlsx")
    wb.save(xlsx_path)

    return {
        "mean": mean_z,
        "var": var_z,
        "alpha": alpha,
        "coverage": (cov1, cov2, cov3),
        "tails": tails,
        "asymmetry": (pos_mean, neg_mean),
    }


def export_uncertainty_results(results, csv_path="uncertainty_comparison_results.csv", sort_by="method", bootstrap=False):
    """
    Convert a list of uncertainty comparison results into a sorted and flattened CSV table.

    Parameters
    ----------
    results : list of dict
        List containing entries with 'filename', 'linespacing', 'method', and nested 'stats'.
    csv_path : str, optional
        Path to save the CSV file (default is 'uncertainty_comparison_results.csv').
    sort_by : str, optional
        Field to sort by, e.g., 'method', 'percentage', or 'rmse'. Default is 'method'.

    Returns
    -------
    pd.DataFrame
        The flattened and sorted DataFrame of results.
    """

    # Optional renaming for nicer output
    method_map = {
        "amplitude": "ASD",
        "psd": "PSD",
        "diff_mean": "SAR",
        "diff_max": "SMR",
        "diff_envelope1": "SER",
    }

    thresholds = (3,5,10)

    # Flatten the results
    flat_results = []
    for r in results:
        stats = r["stats"]
        row = {
            "Seabed": r["filename"].replace(".tif", ""),
            "Method": method_map.get(r["method"], r["method"]),
            "Sampling": r["sampling_method"],
            "Spacing": r["linespacing"],
            "Pass": float(stats["percentage"]),
            "RMSE": float(stats["rmse"]),
            "MAE": float(stats["mae"]),
            "Bias": float(stats["mean"]),
            "Std": float(stats["std_dev"]),
            "Sharpness": float(stats["sharp"]),
            "Correlation": float(stats["corr"]),
            "Severity_mean": float(stats["severity_mean"]),
            "Severity_tail": float(stats["severity_tail"]),
            "Severity_score": float(stats["severity_score"]),
            "Over_mean": float(stats["over_mean"]),
            "Over_tail": float(stats["over_tail"]),
            "Over_penalty": float(stats["over_penalty"]),
            "Reliability_error": float(stats["reliability_error"]),
            "Mean_norm_res": float(stats["mean_norm_res"]),
            "Var_norm_res": float(stats["var_norm_res"]),

            "Total_N": int(stats["Total_N"]),
            "Uncertainty": r["uncertainty"],
            "Residuals": r["residuals"]
        }

        for t in thresholds:
            row.update({
                f"P_gt_{t}": stats[f"P_gt_{t}"],
                f"Gaussian_{t}": stats[f"Gaussian_{t}"],
                f"Ratio_{t}": stats[f"Ratio_{t}"],
                f"Count_{t}": stats[f"Count_{t}"],
            })

        if bootstrap:
            ci_stats = r["ci_stats"]
            row.update({
            "CI_Pass Mean": float(ci_stats["mean"]),
            "CI_Lower": float(ci_stats["ci_lower"]),
            "CI_Upper": float(ci_stats["ci_upper"]),
            "CI_Std": float(ci_stats["std"]),
            "Boot": (ci_stats["boot"]),})

        flat_results.append(row)

    # Convert to DataFrame
    df = pd.DataFrame(flat_results)

    # Sort the DataFrame
    if sort_by in df.columns:
        df = df.sort_values(by=sort_by, ascending=True)
    elif sort_by == "method":
        df = df.sort_values(by="Method", ascending=True)

    # Export to CSV
    df.to_excel(csv_path, index=False)
    print(f"Results exported to: {csv_path}")

    return df, csv_path


